import openpyxl
import pickle, datetime
import os, sqlite3

ScriptDir = pickle.load( open('dir.p','rb'))
EstoqueDir = ScriptDir+'estoque'

#função para adicionar dados na db
def add_estoque(item, qtd, data, cnpj):
    qtd = int(qtd)
    db.execute('INSERT into est_atual (item, qtd, data_att, cnpj) values (?, ?, ?, ?)',
                  (item, qtd, data, cnpj))

#função para abrir o arquivo de excel e pegar os dados específicos de venda
def importa_arquivo_estoque(DIRarquivo):
    wb = openpyxl.load_workbook(DIRarquivo)
    planilha = wb.get_sheet_by_name('Folha1')
    linhamax = planilha.max_row
    QtdImportada = 0

    #loop para adicionar os dados da planilha
    for linha in range(1,linhamax):
        exl_qtd = str(planilha.cell(row=linha, column=10).value)
        try:
            exl_qtd = int(exl_qtd)
            exl_cod = str(planilha.cell(row=linha,column=4).value)
            exl_cnpj = str(planilha.cell(row=26,column=7).value)[0:18]
            data_hj = datetime.date.today()
            if exl_cod != 'None' and exl_qtd != 'None' and isinstance(exl_qtd, int):
                QtdImportada +=1
                add_estoque(exl_cod,exl_qtd,data_hj,exl_cnpj)
        except ValueError:
            continue
    print(str(QtdImportada)+' itens importados!')
#função para renomear arquivo importado, SÓ ENTRA NOME DO ARQUIVO COM A EXTENSÃO
def salva_importado(xlsx):
    DataHoje = str(datetime.date.today())
    HistEst.append((xlsx, DataHoje))
    pickle.dump(HistEst, open(EstoqueDir + '\\est.p', "wb"))

#puxa diretório salvo pelo setup
ScriptDir = pickle.load( open('dir.p','rb'))

#conexão com db
DBDir = ScriptDir+'database'
conn = sqlite3.connect(DBDir+'/database.db')
db = conn.cursor()

#verificação dos arquivos de excel que estão nas pastas de estoque
DocsEstoque = os.listdir(EstoqueDir)
if DocsEstoque == []:
    exit(print('Diretório Vazio'))
else:
    print('Analisando Diretório para verificar arquivos não importados')

#verificação do histórico de arquivos importados
#se não tiver nenhum arquivo é criado início com o excel que tiver
try:
    HistEst = pickle.load( open(EstoqueDir+'\\est.p', 'rb'))
    print('Verificando arquivos na pasta para importação de dados')
    CheckHist = []
    for x in range(len(HistEst)):
        CheckHist.append(HistEst[x][0])

    for x in range(len(DocsEstoque)):
        if DocsEstoque[x].count('.xlsx')>0:
            if CheckHist.count(DocsEstoque[x]) >0:
                print(DocsEstoque[x]+ ' já importado!')
            else:
                print('Iniciar importação do arquivo ' + DocsEstoque[x])
                importa_arquivo_estoque(EstoqueDir + '\\' + DocsEstoque[x])
                salva_importado(DocsEstoque[x])
                print('Arquivo ' + DocsEstoque[x] + ' importado e histórico salvo')
                continue
        else:
            print('Arquivo ' + DocsEstoque[x] + ' não compatível')
            continue

except FileNotFoundError:
    print('Nennhum histórico antecedente, criando estoque inicial')
    HistEst = []
    if len(DocsEstoque) == 1:
        arquivo = EstoqueDir+"\\"+DocsEstoque[0]
        importa_arquivo_estoque(arquivo)
        salva_importado(DocsEstoque[0])
    else:
        print('Vários arquivos encontrados, gerando estoques múltiplos')
        for x in range(len(DocsEstoque)):
            arquivo = EstoqueDir+'\\'+DocsEstoque[x]
            importa_arquivo_estoque(arquivo)
            salva_importado(DocsEstoque[x])
    print('Primeiro acesso salvo')

conn.commit()
conn.close()