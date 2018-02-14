import openpyxl
import pickle, datetime
import os, sqlite3


DateStart = datetime.datetime.now()


#função para adicionar dados na db
def add_vendas(data, item, qtd, valor, cnpj):
    data = data[6:10]+"-"+data[3:5]+"-"+data[0:2]
    qtd = int(qtd)
    db.execute('INSERT into vendas (data, item, qtd, valor, cnpj) values (?, ?, ?, ?, ?)',
                  (data, item, qtd, valor, cnpj))


#função para abrir o arquivo de excel e pegar os dados específicos de venda
def importa_arquivo_venda(DIRarquivo):
    #abre a planilha
    wb = openpyxl.load_workbook(DIRarquivo)
    planilha = wb.get_sheet_by_name('Folha1')

    #Var Statements
    #Essas coordenadas servem para procurar o cabeçalho da planilha e identificar em qual coluna encontrar-se-ão os
    #valores
    ProcuraCoords = {}
    ProcuraCoords['Cód Produto'] = (0,0)
    ProcuraCoords['Data'] = (0,0)
    ProcuraCoords['Qtd.'] = (0,0)
    ProcuraCoords['Vl Unit.'] = (0,0)

    #Range max para procurar, não sei se é necessário, mas enfim
    linhamax = planilha.max_row
    colunamax = planilha.max_column
    QtdImportada = 0

    #código bandaid para não me foder na procura de cnpj
    CNPJ1 = str(planilha.cell(row=22, column=6).value)[0:18]
    CNPJ2 = str(planilha.cell(row=22, column=5).value)[0:18]
    if CNPJ1 == 'None':
        CNPJReal = CNPJ2
    else:
        CNPJReal = CNPJ1

    #loop para adicionar os dados da planilha
    #primeiro procuro as coordendas que coloquei acima
    for linha in range(1,linhamax):
        for coluna in range(1,colunamax):
            celula = str(planilha.cell(row=linha, column=coluna).value)
            if list(ProcuraCoords.keys()).count(celula) > 0:
                ProcuraCoords[celula] = (linha,coluna)
                if list(ProcuraCoords.values()).count((0,0)) == 0:
                    break
        if list(ProcuraCoords.values()).count((0, 0)) == 0:
            break
    LinhaIn = list(ProcuraCoords.values())[0][0]+1
    CodCoords, DataCoords = ProcuraCoords['Cód Produto'][1], ProcuraCoords['Data'][1]
    QtdCoords, VlrCoords = ProcuraCoords['Qtd.'][1], ProcuraCoords['Vl Unit.'][1]

    #após as coords encontradas eu faço outro loop para adicionar os valores, hope it works
    for linha in range(LinhaIn,linhamax):
        CodValor = str(planilha.cell(row=linha,column=CodCoords).value)
        DataValor = str(planilha.cell(row=linha,column=DataCoords).value)
        QtdValor = str(planilha.cell(row=linha,column=QtdCoords).value)
        VlrValor = str(planilha.cell(row=linha,column=VlrCoords).value)
        if len(DataValor) == 10 and DataValor.count('/') == 2:
            add_vendas(DataValor,CodValor,QtdValor,VlrValor,CNPJReal)
            QtdImportada += 1
    print(DIRarquivo, 'teve', QtdImportada, 'registros importados.', sep=' ')


#função para renomear arquivo importado, SÓ ENTRA NOME DO ARQUIVO COM A EXTENSÃO
def salva_importado(xlsx):
    DataHoje = datetime.date.today()
    HistVendas.append((xlsx, DataHoje))
    pickle.dump(HistVendas, open(VendasDir + '\\vendas.p', "wb"))


#puxa diretório salvo pelo setup
ScriptDir = pickle.load( open('dir.p','rb'))

#conexão com db
DBDir = ScriptDir+'database'
conn = sqlite3.connect(DBDir+'/database.db')
db = conn.cursor()

#definição das pastas de vendas e estoque
VendasDir = ScriptDir+'vendas'

#verificação dos arquivos de excel que estão nas pastas de venda
DocsVenda = os.listdir(VendasDir)
if DocsVenda == []:
    exit(print('Diretório Vazio'))
else:
    print('Analisando Diretório para verificar arquivos não importados')

#verificação do histórico de arquivos importados
#se não tiver nenhum arquivo é criado início com o excel que tiver
try:
    HistVendas = pickle.load( open(VendasDir+'\\vendas.p', 'rb'))
    print('Verificando arquivos na pasta para importação de dados')
    ArquivosHist = []
    for x in range(len(HistVendas)):
        ArquivosHist.append(HistVendas[x][0])
    for x in range(len(DocsVenda)):
        if ArquivosHist.count(DocsVenda[x]) == 0 and DocsVenda[x].count('.xlsx') > 0:
            print('Importando arquivo',DocsVenda[x],sep=' ')
            arquivo = VendasDir+'\\'+DocsVenda[x]
            importa_arquivo_venda(arquivo)
            salva_importado(DocsVenda[x])
        elif ArquivosHist.count(DocsVenda[x]) != 0:
            print('Arquivo',DocsVenda[x],'já importado.',sep=' ')
        else:
            print('Arquivo',DocsVenda[x],'não compartível',sep=' ')


except FileNotFoundError:
    print('Nennhum histórico antecedente, criando vendas iniciais')
    HistVendas = []
    if len(DocsVenda) == 1:
        arquivo = VendasDir+'\\'+DocsVenda[0]
        importa_arquivo_venda(arquivo)
        salva_importado(DocsVenda[0])
    else:
        print('Vários arquivos encontrados, gerando estoques múltiplos')
        for x in range(len(DocsVenda)):
            arquivo = VendasDir + '\\' + DocsVenda[x]
            importa_arquivo_venda(arquivo)
            salva_importado(DocsVenda[x])
    print('Primeiro acesso salvo')


conn.commit()
conn.close()
